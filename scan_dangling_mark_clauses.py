import openpyxl
import spacy
from collections import defaultdict


def main() -> None:
    wb = openpyxl.load_workbook("interview_clauses.xlsx", read_only=True)
    ws = wb.active

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(header)}
    required = ["Filename", "Sent", "Clause", "Speaker", "Text"]
    missing = [c for c in required if c not in col]
    if missing:
        raise SystemExit(f"Missing columns in Excel: {missing}. Found: {header}")

    metas: list[tuple[str, int, int, str, str]] = []
    texts: list[str] = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        text = (r[col["Text"]] or "").strip()
        if not text:
            continue
        metas.append(
            (
                r[col["Filename"]],
                int(r[col["Sent"]]),
                int(r[col["Clause"]]),
                r[col["Speaker"]],
                text,
            )
        )
        texts.append(text)

    nlp = spacy.load("en_core_web_lg")

    hits: list[tuple[str, str, int, int, str, str, str, str, str]] = []

    for meta, doc in zip(metas, nlp.pipe(texts, batch_size=64)):
        tokens = [t for t in doc if not t.is_space and not t.is_punct]
        if len(tokens) != 1:
            continue
        tok = tokens[0]
        if tok.dep_ != "mark":
            continue

        fn, sent, clause, speaker, text = meta
        hits.append(
            (
                tok.text.lower(),
                fn,
                sent,
                clause,
                speaker,
                text,
                tok.text,
                tok.pos_,
                tok.tag_,
            )
        )

    print(f"Scanned clauses: {len(texts)}")
    print(f"Dangling one-word mark clauses: {len(hits)}")

    by_word: dict[str, list[tuple[str, str, int, int, str, str, str, str, str]]] = defaultdict(list)
    for h in hits:
        by_word[h[0]].append(h)

    for word in sorted(by_word, key=lambda w: (-len(by_word[w]), w)):
        group = by_word[word]
        print(f"\n{word} ({len(group)})")
        for (_, fn, sent, clause, speaker, _text, tok_text, pos, tag) in group[:15]:
            print(f"  {fn} | {speaker} | {sent}.{clause} | {tok_text}/{pos}/{tag}")

    if hits:
        print("\nExamples (first 25):")
        for (_w, fn, sent, clause, speaker, text, _tok_text, _pos, _tag) in hits[:25]:
            print(f"- {fn} | {speaker} | {sent}.{clause} | {text}")


if __name__ == "__main__":
    main()
